import streamlit as st
import pandas as pd
import json
import re
import io
import base64
import google.generativeai as genai
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# --- הגדרות דף ---
st.set_page_config(
    page_title="PC Pro Manager",
    page_icon="💻",
    layout="wide",
    initial_sidebar_state="expanded",
)

# --- עיצוב CSS ירוק ניאון ---
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&family=Syne:wght@400;700;800&display=swap');

:root {
    --bg-primary: #0a0f0a;
    --bg-secondary: #0f1a0f;
    --bg-card: #111c11;
    --green-neon: #4dbb4d;
    --green-bright: #3a9c3a;
    --green-dim: #1a3a1a;
    --text-main: #e8f5e8;
    --border: #1e3a1e;
}

html, body, [class*="css"] {
    font-family: 'Syne', sans-serif !important;
    background-color: var(--bg-primary) !important;
    color: var(--text-main) !important;
    direction: rtl;
}

.stApp {
    background: linear-gradient(135deg, #0a0f0a 0%, #0d160d 50%, #0a0f0a 100%) !important;
}

.main-header {
    background: linear-gradient(90deg, var(--bg-card) 0%, var(--green-dim) 100%);
    border: 1px solid var(--green-bright);
    border-radius: 15px;
    padding: 20px;
    margin-bottom: 25px;
    text-align: center;
}

.inventory-item-box {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 15px;
    margin-bottom: 10px;
}

.available-qty {
    color: var(--green-neon);
    font-size: 1.2rem;
    font-weight: 800;
}

.price-tag {
    font-family: 'JetBrains Mono', monospace;
    color: #ffd700;
    font-size: 1.3rem;
}

.cart-item-row {
    background: rgba(255,255,255,0.03);
    border-right: 4px solid var(--green-neon);
    padding: 10px;
    margin-bottom: 8px;
    border-radius: 4px;
}
</style>
""", unsafe_allow_html=True)

# --- ניהול State ---
if "inventory" not in st.session_state: st.session_state.inventory = None
if "cart" not in st.session_state: st.session_state.cart = []
if "extracted_items" not in st.session_state: st.session_state.extracted_items = []

# --- פונקציות עזר ---
def format_num(val):
    try:
        if pd.isna(val): return 0.0
        if isinstance(val, (int, float)): return float(val)
        cleaned = re.sub(r'[^\d.-]', '', str(val))
        return float(cleaned) if cleaned else 0.0
    except: return 0.0

def find_column(df, possible_names):
    for col in df.columns:
        clean_col = str(col).strip().replace(" ", "").replace("'", "").replace('"', "")
        for target in possible_names:
            if clean_col == target.replace(" ", "").replace("'", "").replace('"', ""):
                return col
    return None

def parse_clean_json(text):
    text = re.sub(r"```json|```", "", text, flags=re.IGNORECASE).strip()
    match = re.search(r"(\[.*\]|\{.*\})", text, re.DOTALL)
    return match.group(1) if match else text

def export_to_excel(cart_data, margin):
    wb = Workbook()
    ws = wb.active
    ws.title = "הצעת מחיר"
    ws.sheet_view.rightToLeft = True
    headers = ["תיאור פריט", "מק\"ט", "כמות", "מחיר ללקוח", "סה\"כ"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="1A3A1A", end_color="1A3A1A", fill_type="solid")
        cell.font = Font(bold=True, color="4DBB4D")
    for item in cart_data:
        p_cust = round(format_num(item.get('price', 0)) * (1 + margin/100), 2)
        qty = int(item.get('quantity', 1))
        ws.append([item.get('description', ''), item.get('sku', ''), qty, p_cust, p_cust * qty])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- Sidebar ---
with st.sidebar:
    st.markdown("### ⚙️ הגדרות מערכת")
    api_key = st.text_input("Gemini API Key", type="password")
    profit_margin = st.slider("רווח ללקוח (%)", 0, 100, 20)
    
    model = None
    if api_key:
        try:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-1.5-flash')
            st.success("AI מוכן לפעולה")
        except: st.error("Key Error")

# --- דף ראשי ---
st.markdown('<div class="main-header"><h1>PC Pro Manager</h1></div>', unsafe_allow_html=True)

tabs = st.tabs(["🔍 חיפוש במלאי", "📝 חילוץ ממייל/תמונה", "🛒 סל והצעת מחיר"])

with tabs[0]:
    c_up1, c_up2 = st.columns([1, 2])
    with c_up1:
        uploaded_inventory = st.file_uploader("טען אקסל מלאי משפחה מרוכזת", type=["xlsx"])
        if uploaded_inventory:
            try:
                df = pd.read_excel(uploaded_inventory, engine='openpyxl')
                st.session_state.inventory = df
                st.success(f"נטענו {len(df)} פריטים")
            except: st.error("שגיאה בטעינה")

    with c_up2:
        search_query = st.text_input("חפש מוצר (מק\"ט / תיאור):")

    if st.session_state.inventory is not None and search_query:
        df = st.session_state.inventory
        col_sku = find_column(df, ["מק'ט", "מקט", "מק\"ט"])
        col_desc = find_column(df, ["תאור מוצר", "תיאור מוצר"])
        col_bal = find_column(df, ["יתרה מחסני מכירה", "יתרה"])
        col_orders = find_column(df, ["הזמנות לקוח"])
        col_purch = find_column(df, ["כמות ברכש"])
        col_price = find_column(df, ["מחיר מוצג לסוכן $", "מחיר סוכן"])
        
        res = df[df.apply(lambda r: search_query.lower() in str(r).lower(), axis=1)]
        
        if not res.empty:
            for idx, row in res.iterrows():
                sku = str(row[col_sku]) if col_sku else "N/A"
                desc = str(row[col_desc]) if col_desc else "N/A"
                v_bal = format_num(row[col_bal]) if col_bal else 0
                v_orders = format_num(row[col_orders]) if col_orders else 0
                v_purch = format_num(row[col_purch]) if col_purch else 0
                v_price = format_num(row[col_price]) if col_price else 0
                v_total = v_bal - v_orders + v_purch
                
                st.markdown(f"""
                <div class="inventory-item-box">
                    <div style="display: flex; justify-content: space-between;">
                        <div><b>{desc}</b><br><small>מק"ט: {sku}</small></div>
                        <div class="price-tag">${v_price}</div>
                    </div>
                    <div style="display: flex; gap: 15px; margin-top: 10px; border-top: 1px solid #1e3a1e; padding-top: 10px;">
                        <div style="flex:1"><small>יתרה:</small><br><b>{int(v_bal)}</b></div>
                        <div style="flex:1"><small>הזמנות:</small><br><b style="color:#ff4b4b">-{int(v_orders)}</b></div>
                        <div style="flex:1"><small>ברכש:</small><br><b style="color:#4d88ff">+{int(v_purch)}</b></div>
                        <div style="flex:1; background:rgba(77,187,77,0.1); border-radius:5px; text-align:center;">
                            <small>זמין:</small><br><b class="available-qty">{int(v_total)}</b>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                if st.button(f"➕ הוסף להצעה", key=f"add_{idx}"):
                    st.session_state.cart.append({"description": desc, "sku": sku, "price": v_price, "quantity": 1})
                    st.toast(f"התווסף {sku}")

with tabs[1]:
    st.subheader("חילוץ מוצרים חכם (AI)")
    st.info("הדבק כאן טקסט ממייל או העלה צילום מסך של הצעת מחיר מספק.")
    
    c_ai1, c_ai2 = st.columns(2)
    with c_ai1:
        mail_input = st.text_area("הדבק טקסט (מייל/רשימה):", height=250, placeholder="הדבק כאן את תוכן המייל...")
    with c_ai2:
        img_input = st.file_uploader("או העלה צילום מסך:", type=["png", "jpg", "jpeg"])

    if st.button("🔍 נתח וחלץ נתונים") and model:
        with st.spinner("AI מנתח..."):
            prompt = """
            You are a professional computer hardware inventory assistant. 
            Extract all products from the provided text or image.
            Return ONLY a JSON list of objects with these exact keys: 'description', 'sku', 'quantity', 'price'.
            Rules:
            1. 'price' must be a clean number (remove $ or symbols).
            2. 'sku' is the part number/manufacturer code.
            3. 'description' is the full product name.
            4. If a value is missing, use an empty string.
            5. Search carefully for quantities and prices.
            """
            
            if img_input:
                img_data = base64.b64encode(img_input.read()).decode()
                response = model.generate_content([prompt, {"inline_data": {"mime_type": "image/jpeg", "data": img_data}}])
            else:
                response = model.generate_content(f"{prompt}\n\nContent to analyze:\n{mail_input}")
            
            try:
                st.session_state.extracted_items = json.loads(parse_clean_json(response.text))
                st.success(f"נמצאו {len(st.session_state.extracted_items)} פריטים!")
            except:
                st.error("לא הצלחתי לחלץ נתונים באופן אוטומטי. נסה טקסט ברור יותר או צילום מסך נקי.")

    if st.session_state.extracted_items:
        st.write("### פריטים שזוהו:")
        for idx, item in enumerate(st.session_state.extracted_items):
            c1, c2, c3, c4, c5 = st.columns([3, 2, 1, 1, 1])
            with c1: st.text_input("תיאור", item.get('description'), key=f"ex_desc_{idx}")
            with c2: st.text_input("מק\"ט", item.get('sku'), key=f"ex_sku_{idx}")
            with c3: st.number_input("כמות", value=int(item.get('quantity', 1)) or 1, key=f"ex_qty_{idx}")
            with c4: st.number_input("מחיר", value=float(item.get('price', 0)), key=f"ex_prc_{idx}")
            with c5: 
                if st.button("➕", key=f"ex_add_{idx}"):
                    st.session_state.cart.append({
                        "description": item.get('description'),
                        "sku": item.get('sku'),
                        "price": item.get('price'),
                        "quantity": item.get('quantity', 1)
                    })
                    st.toast("נוסף לסל")
        
        if st.button("✅ הוסף את כל הפריטים לסל"):
            for item in st.session_state.extracted_items:
                st.session_state.cart.append(item)
            st.session_state.extracted_items = []
            st.success("הכל נוסף לסל!")
            st.rerun()

with tabs[2]:
    st.subheader("🛒 ריכוז הצעת מחיר סופית")
    if not st.session_state.cart:
        st.info("הסל ריק")
    else:
        total_sum = 0
        for i, item in enumerate(st.session_state.cart):
            p_agent = format_num(item.get('price', 0))
            p_cust = round(p_agent * (1 + profit_margin/100), 2)
            qty = int(item.get('quantity', 1))
            total_sum += (p_cust * qty)
            
            c_info, c_del = st.columns([6, 1])
            with c_info:
                st.markdown(f"""
                <div class="cart-item-row">
                    <div style="display:flex; justify-content:space-between">
                        <b>{item['description']}</b>
                        <span style="color:var(--green-neon)">${p_cust} x {qty}</span>
                    </div>
                    <small>מק"ט: {item['sku']}</small>
                </div>
                """, unsafe_allow_html=True)
            with c_del:
                if st.button("🗑️", key=f"del_final_{i}"):
                    st.session_state.cart.pop(i)
                    st.rerun()
        
        st.markdown(f"### סה\"כ הצעת מחיר: **${round(total_sum, 2)}**")
        st.divider()
        
        c_btn1, c_btn2, c_btn3 = st.columns(3)
        with c_btn1:
            if st.button("🗑️ נקה הכל"):
                st.session_state.cart = []
                st.rerun()
        with c_btn2:
            xl_data = export_to_excel(st.session_state.cart, profit_margin)
            st.download_button("📥 הורד קובץ Excel", data=xl_data, file_name="PC_Pro_Quote.xlsx")
        with c_btn3:
            st.write("להדפסה ל-PDF לחץ במקלדת:")
            st.code("Ctrl + P")
